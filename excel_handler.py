import csv
import os

import numpy as np
import openpyxl
import pandas as pd

from configuration import MaterialProperties, LoadCase, Panel, PanelLayers, ReserveFactors, Stringer, IO


def read_excel_template(file_path, material_properties: MaterialProperties, start_row=7, end_row=9, column=2):
    for root, dirs, files in os.walk("./input"):
        for file in files:
            # Get the full path of the file
            _file_path = os.path.join(root, file)
            fill_csv_with_commas(_file_path)

    # Read the CSV file
    df = pd.read_csv(file_path, header=None)  # header=None to treat all rows as data

    # Extract the specific rows and column
    specific_rows = df.iloc[start_row - 1:end_row, column - 1]  # Subtract 1 to convert to zero-based indexing
    material_properties.E1 = specific_rows[0]
    material_properties.E2 = specific_rows[1]
    material_properties.G12 = specific_rows[2]
    material_properties.nu12 = 0.33
    material_properties.R1t = 3050
    material_properties.R1c = 1500
    material_properties.R2t = 300
    material_properties.R2c = 50
    material_properties.R21 = 100


def parse_constants(E_hom_flange, E_hom_web, E_hom_skin, EI, z_EC_comb, r_gyr, lamda, lamda_crit):
    io = IO()
    wb = openpyxl.load_workbook(io.output_file)
    ws = wb.get_sheet_by_name(io.sheet_name_output)
    column = 2
    for row in range(69, 73):
        cellref = ws.cell(row=row, column=column)
        cellref.value = E_hom_flange
        cellref = ws.cell(row=row, column=column + 1)
        cellref.value = E_hom_web
        cellref = ws.cell(row=row, column=column + 2)
        cellref.value = E_hom_skin
        cellref = ws.cell(row=row, column=column + 3)
        cellref.value = E_hom_skin
        cellref = ws.cell(row=row, column=column + 4)
        cellref.value = z_EC_comb
        cellref = ws.cell(row=row, column=column + 5)
        cellref.value = EI
        cellref = ws.cell(row=row, column=column + 6)
        cellref.value = r_gyr
        cellref = ws.cell(row=row, column=column + 7)
        cellref.value = lamda
        cellref = ws.cell(row=row, column=column + 8)
        cellref.value = lamda_crit
    wb.save(io.output_file)


def read_excel_input(file_path="./input"):
    df_xx = pd.read_csv(file_path + "/panel_xx.csv", header=None, delimiter=',', skiprows=11)
    df_yy = pd.read_csv(file_path + "/panel_yy.csv", header=None, delimiter=',', skiprows=11)
    df_xy = pd.read_csv(file_path + "/panel_xy.csv", header=None, delimiter=',', skiprows=11)
    df_stress = pd.read_csv(file_path + "/panel_stress.csv", header=None, delimiter=',', skiprows=11)
    df_str_strain = pd.read_csv(file_path + "/stringer_strain.csv", header=None, delimiter=',', skiprows=11)
    df_str_stress = pd.read_csv(file_path + "/stringer_stress.csv", header=None, delimiter=',', skiprows=11)
    load_cases = []
    index_first = 0
    index_second = 0
    index_third = 0
    for i in range(3):
        panel_layers = []
        panels = []
        stringers = []
        while index_first < len(df_xx) and int(df_xx.loc[index_first, 2]) == i + 1:
            e_id = int(df_xx.loc[index_first, 0])
            layer = int(df_xx.loc[index_first, 4].replace('Ply', '').strip())
            sig_xx = float(df_xx.loc[index_first, 5])
            sig_yy = float(df_yy.loc[index_first, 5])
            sig_xy = float(df_xy.loc[index_first, 5])
            panel_layer = PanelLayers(e_id=e_id, layer=layer, sig_xx=sig_xx, sig_yy=sig_yy, sig_xy=sig_xy)
            panel_layers.append(panel_layer)
            index_first += 1
        while index_second < len(df_stress) and int(df_stress.loc[index_second, 2]) == i + 1:
            panel = Panel(
                e_id=int(df_stress.loc[index_second, 0]),
                sigma_1=float(df_stress.loc[index_second, 5]),
                sigma_2=float(df_stress.loc[index_second, 7]),
                tau=float(df_stress.loc[index_second, 6])
            )
            panels.append(panel)
            index_second += 1
        while index_third < len(df_str_strain) and int(df_str_strain.loc[index_third, 2]) == i + 1:
            stringer = Stringer(
                e_id=int(df_str_strain.loc[index_third, 0]),
                strain=float(df_str_strain.loc[index_third, 4]),
                stress=float(df_str_stress.loc[index_third, 4])
            )
            stringers.append(stringer)
            index_third += 1
        load_cases.append(LoadCase(i, panel_layers, panels, stringers))
    return load_cases


def fill_csv_with_commas(input_file):
    with open(input_file, 'r') as infile:
        reader = csv.reader(infile)
        rows = list(reader)
    max_delimiters = max(len(row) - 1 for row in rows)
    for row in rows:
        num_commas_needed = max_delimiters - (len(row) - 1)
        row.extend([''] * num_commas_needed)
    with open(input_file, 'w', newline='') as outfile:
        writer = csv.writer(outfile)
        writer.writerows(rows)


def write_excel_template(load_cases: [LoadCase], io: IO = IO()):
    wb = openpyxl.load_workbook(io.output_file)
    ws = wb.get_sheet_by_name(io.sheet_name_output)
    parse_reserve_factors(ws, load_cases)
    parse_plane_analysis(ws, load_cases)
    parse_stringer_analysis(ws, load_cases)
    wb.save(io.output_file)


def parse_ADB_matrix(A, B, D, io: IO = IO()):
    wb = openpyxl.load_workbook(io.output_file)
    ws = wb.get_sheet_by_name(io.sheet_name_output)
    parse_matrix(A, ws, 17)
    parse_matrix(B, ws, 21)
    parse_matrix(D, ws, 25)
    wb.save(io.output_file)


def parse_matrix(mat, ws, start_row):
    for i, row in enumerate(mat):
        for j, element in enumerate(row):
            ws.cell(row=i + start_row, column=j + 1).value = element


def parse_reserve_factors(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row_loads = 32
        index = 0
        for k, load in enumerate(load_case.PanelsLayers):
            if load.e_id == 1:
                cellref = ws.cell(row=row_loads + index, column=column)
                cellref.value = load.reserve_factor.RF_FF
                cellref = ws.cell(row=row_loads + index, column=column + 1)
                cellref.value = load.reserve_factor.RF_IFF
                cellref = ws.cell(row=row_loads + index, column=column + 2)
                cellref.value = load.mode
                cellref = ws.cell(row=row_loads + index, column=column + 3)
                cellref.value = load.reserve_factor.RF_strength
                index += 1
        row_loads = 40
        for k, load in enumerate(load_case.Stringers):
            if load.e_id == 40:
                cellref = ws.cell(row=row_loads + k, column=column)
                cellref.value = load.reserve_factor.RF_FF
                cellref = ws.cell(row=row_loads + k, column=column + 1)
                cellref.value = load.reserve_factor.RF_IFF
                cellref = ws.cell(row=row_loads + k, column=column + 2)
                cellref.value = load.mode
                cellref = ws.cell(row=row_loads + k, column=column + 3)
                cellref.value = load.reserve_factor.RF_strength
            else:
                break
        column += 6


def parse_stringer_strength(stringers: [Stringer], column):
    io = IO()
    wb = openpyxl.load_workbook(io.output_file)
    ws = wb.get_sheet_by_name(io.sheet_name_output)
    row_loads = 40
    for k, load in enumerate(stringers):
        cellref = ws.cell(row=row_loads + k, column=column)
        cellref.value = load.reserve_factor.RF_FF
        cellref = ws.cell(row=row_loads + k, column=column + 1)
        cellref.value = load.reserve_factor.RF_IFF
        cellref = ws.cell(row=row_loads + k, column=column + 2)
        cellref.value = load.mode
        cellref = ws.cell(row=row_loads + k, column=column + 3)
        cellref.value = load.reserve_factor.RF_strength
    wb.save(io.output_file)


def parse_plane_analysis(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row = 53
        for k, panel in enumerate(load_case.Panels[:5]):
            cellref = ws.cell(row=row + k, column=column)
            cellref.value = panel.sig_xx_avg
            cellref = ws.cell(row=row + k, column=column + 1)
            cellref.value = panel.sig_yy_avg
            cellref = ws.cell(row=row + k, column=column + 2)
            cellref.value = panel.sig_xy_avg
            cellref = ws.cell(row=row + k, column=column + 3)
            cellref.value = panel.sig_crit_shear
            cellref = ws.cell(row=row + k, column=column + 4)
            cellref.value = panel.sig_crit_biax
            cellref = ws.cell(row=row + k, column=column + 5)
            cellref.value = panel.RF_panel_buckling
        column += 8


def parse_stringer_analysis(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row = 62
        row_geo = 68
        for k, stringer in enumerate(load_case.Stringers[:4]):
            cellref = ws.cell(row=row + k, column=column)
            cellref.value = stringer.sig_combined
            cellref = ws.cell(row=row + k, column=column + 1)
            cellref.value = stringer.sig_crip
            cellref = ws.cell(row=row + k, column=column + 2)
            cellref.value = stringer.RF_combined_buckling
            # parse_geometric_properties(ws, stringer, row_geo + k)
        column += 5


def parse_geometric_properties(ws, stringer: Stringer, row):
    column = 2
    for k, geo in enumerate(stringer.geometrics):
        cellref = ws.cell(row=row, column=column)
        cellref.value = geo.I
        cellref = ws.cell(row=row, column=column + 1)
        cellref.value = geo.r_gyr
        cellref = ws.cell(row=row, column=column + 2)
        cellref.value = geo.lamda
        cellref = ws.cell(row=row, column=column + 3)
        cellref.value = geo.lamda_crit

# read_excel_template("./input/ASE_Project2024_task2_1_Template_3771075.csv", None)
